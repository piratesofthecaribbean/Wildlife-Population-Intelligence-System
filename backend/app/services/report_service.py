"""
Wildlife Monitoring Reports — PDF and Excel export.
"""

import io
from datetime import datetime, timezone
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


class ReportService:
    @staticmethod
    def generate_monitoring_report(metrics: Dict[str, Any], history: list) -> bytes:
        """Generate a PDF wildlife monitoring report."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(
            "Wildlife Population Intelligence System",
            styles["Title"],
        ))
        elements.append(Paragraph(
            "Wildlife Monitoring Report",
            styles["Heading2"],
        ))
        elements.append(Paragraph(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            styles["Normal"],
        ))
        elements.append(Spacer(1, 20))

        # Summary table
        summary_data = [
            ["Metric", "Value"],
            ["Total Observations", str(metrics.get("total_observations", 0))],
            ["Species Richness", str(metrics.get("species_richness", 0))],
            ["Shannon Diversity Index", str(metrics.get("shannon_diversity_index", 0))],
            ["Simpson Diversity Index", str(metrics.get("simpson_diversity_index", 0))],
            ["Biodiversity Index", f"{metrics.get('biodiversity_index', 0)}%"],
            ["Biodiversity Health", metrics.get("biodiversity_health", "N/A")],
            ["Endangered Detections", str(metrics.get("endangered_detections", 0))],
            ["Image Observations", str(metrics.get("image_observations", 0))],
            ["Audio Observations", str(metrics.get("audio_observations", 0))],
        ]

        table = Table(summary_data, colWidths=[250, 200])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#166534")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0fdf4")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("PADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))

        # Recent observations
        elements.append(Paragraph("Recent Observations", styles["Heading3"]))
        elements.append(Spacer(1, 10))

        if history:
            obs_data = [["Type", "Species", "Confidence", "Date"]]
            for obs in history[:15]:
                obs_data.append([
                    obs.get("source_type", ""),
                    obs.get("species_name", ""),
                    f"{round(obs.get('confidence', 0) * 100)}%",
                    (obs.get("created_at") or "")[:10],
                ])
            obs_table = Table(obs_data, colWidths=[60, 180, 80, 100])
            obs_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#166534")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]))
            elements.append(obs_table)
        else:
            elements.append(Paragraph("No observations recorded yet.", styles["Normal"]))

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_excel_report(metrics: Dict[str, Any], history: list) -> bytes:
        """Generate an Excel biodiversity report."""
        wb = Workbook()

        # Summary sheet
        ws = wb.active
        ws.title = "Biodiversity Summary"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")

        ws.append(["Wildlife Monitoring Report"])
        ws.append([f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"])
        ws.append([])
        ws.append(["Metric", "Value"])
        for cell in ws[4]:
            cell.font = header_font
            cell.fill = header_fill

        summary_rows = [
            ("Total Observations", metrics.get("total_observations", 0)),
            ("Species Richness", metrics.get("species_richness", 0)),
            ("Shannon Diversity Index", metrics.get("shannon_diversity_index", 0)),
            ("Simpson Diversity Index", metrics.get("simpson_diversity_index", 0)),
            ("Biodiversity Index (%)", metrics.get("biodiversity_index", 0)),
            ("Biodiversity Health", metrics.get("biodiversity_health", "N/A")),
            ("Endangered Detections", metrics.get("endangered_detections", 0)),
        ]
        for row in summary_rows:
            ws.append(list(row))

        # Species distribution sheet
        ws2 = wb.create_sheet("Species Distribution")
        ws2.append(["Species", "Count"])
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
        for item in metrics.get("species_distribution", []):
            ws2.append([item["species"], item["count"]])

        # Observations sheet
        ws3 = wb.create_sheet("Observations")
        ws3.append(["Type", "Species", "Scientific Name", "Confidence", "Endangered", "Date"])
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill
        for obs in history:
            ws3.append([
                obs.get("source_type", ""),
                obs.get("species_name", ""),
                obs.get("scientific_name", ""),
                round(obs.get("confidence", 0), 3),
                "Yes" if obs.get("is_endangered") else "No",
                (obs.get("created_at") or "")[:19],
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
